from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

app = FastAPI()

# ============================================================================
# UTILITY FUNCTIONS - Reusable across different reports
# ============================================================================


def get_severity_color(level: str) -> str:
    """Get background color based on severity level"""
    colors = {
        "high": "FFEEEE",
        "medium": "FFD580",
        "low": "FFFFE0",
        "critical": "FF0000",
    }
    return colors.get(level.lower() if level else "", "FFFFFF")


def get_border_style() -> Border:
    """Get standard border style for cells"""
    thin_border = Side(style="thin", color="00000000")
    return Border(
        top=thin_border, left=thin_border, bottom=thin_border, right=thin_border
    )


def get_alignment(horizontal="center", vertical="middle", wrap_text=True) -> Alignment:
    """Get cell alignment style"""
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)


def get_font(font_type: str = "data") -> Font:
    """Get font style based on type"""
    fonts = {
        "data": Font(bold=False, size=12),
        "resource-header": Font(bold=True, size=14),
        "header": Font(bold=True, size=16),
    }
    return fonts.get(font_type, Font(bold=True, size=16))


def get_fill(color: str) -> PatternFill:
    """Get cell fill pattern with color"""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def format_date(date_value: Any) -> str:
    """Format date to readable string"""
    if isinstance(date_value, datetime):
        return date_value.strftime("%Y-%m-%d %H:%M:%S")
    return str(date_value) if date_value else "N/A"


def apply_cell_style(
    cell, value, font_type="data", bg_color="FFFFFF", border=True, alignment=True
):
    """Apply comprehensive styling to a cell"""
    cell.value = value
    cell.font = get_font(font_type)
    cell.fill = get_fill(bg_color)
    if border:
        cell.border = get_border_style()
    if alignment:
        cell.alignment = get_alignment()


def sort_by_severity(items: List[Dict], key: str = "severity_score") -> List[Dict]:
    """Sort items by severity score in descending order"""
    return sorted(items, key=lambda x: x.get(key, 0), reverse=True)


# ============================================================================
# SUMMARY SHEET GENERATION
# ============================================================================


def create_summary_sheet(    workbook: Workbook, data):
    """Create the summary worksheet with account and scan information"""
    sheet = workbook.create_sheet("Summary", 0)

    # Set column widths
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 30

    # Summary fields
    summary_data = [
        {"field": "Name", "value": meta.get("account_name", "")},
        {"field": "Account ID", "value": meta.get("account_id", "")},
        {
            "field": "Last Scanned",
            "value": format_date(meta.get("last_scanned") or meta.get("timestamp")),
        },
        {
            "field": "Regions Scanned",
            "value": (
                ", ".join(meta.get("regions", []))
                if isinstance(meta.get("regions"), list)
                else "N/A"
            ),
        },
        {
            "field": "Services Scanned",
            "value": (
                ", ".join(
                    meta.get("scanned_meta_data", [{}])[0]
                    .get("data", {})
                    .get("services_scanned", [])
                )
                if meta.get("scanned_meta_data")
                else "N/A"
            ),
        },
    ]

    # Insert summary fields
    for i, item in enumerate(summary_data):
        row_index = i + 3
        apply_cell_style(
            sheet[f"B{row_index}"],
            item["field"],
            font_type="resource-header",
            bg_color="FFF2F2F2",
        )
        apply_cell_style(
            sheet[f"C{row_index}"],
            item["value"],
            font_type="resource-header",
            bg_color="FFFFFFFF",
        )

    # Insert scanned metadata table
    meta_start_row = len(summary_data) + 5
    add_scanned_metadata_table(sheet, meta, meta_start_row)

    # Insert security services scan results
    services_start_row = meta_start_row + len(meta.get("scanned_meta_data", [])) * 7 + 2
    add_security_services_table(sheet, security_services, services_start_row)


def add_scanned_metadata_table(sheet, meta: Dict, start_row: int) -> None:
    """Add scanned metadata table to summary sheet"""
    # Header row
    headers = ["Region", "Metric", "Value"]
    for col_idx, header in enumerate(headers, start=2):
        cell = sheet.cell(row=start_row, column=col_idx)
        apply_cell_style(cell, header, font_type="data", bg_color="FFCCE5FF")

    current_row = start_row + 1

    for item in meta.get("scanned_meta_data", []):
        region_start = current_row

        # Metrics
        metrics = [
            {
                "metric": "Total no. of Resources",
                "value": item.get("data", {}).get("total_scanned", 0),
            },
            {
                "metric": "No. of Issues Found",
                "value": item.get("data", {}).get("affected", 0),
            },
            {"metric": "High", "value": item.get("data", {}).get("High", 0)},
            {"metric": "Medium", "value": item.get("data", {}).get("Medium", 0)},
            {"metric": "Low", "value": item.get("data", {}).get("Low", 0)},
            {"metric": "Critical", "value": item.get("data", {}).get("Critical", 0)},
        ]

        # Merge region cells
        sheet.merge_cells(f"B{region_start}:B{region_start + len(metrics) - 1}")
        region_cell = sheet[f"B{region_start}"]
        apply_cell_style(
            region_cell,
            item.get("region", ""),
            font_type="resource-header",
            bg_color="FFFFFF",
        )

        # Add metrics
        for j, entry in enumerate(metrics):
            row = region_start + j
            apply_cell_style(sheet.cell(row=row, column=3), entry["metric"])
            apply_cell_style(sheet.cell(row=row, column=4), entry["value"])

        current_row += len(metrics)


def add_security_services_table(
    sheet, security_services: List[Dict], start_row: int
) -> None:
    """Add security services scan results table"""
    # Title
    title_cell = sheet[f"B{start_row}"]
    apply_cell_style(
        title_cell,
        "Security Services Scan Results",
        font_type="resource-header",
        bg_color="FFFFFF",
    )

    # Headers
    header_row = start_row + 1
    headers = ["Region", "Service", "Enabled", "Recommendation"]
    for col_idx, header in enumerate(headers, start=2):
        cell = sheet.cell(row=header_row, column=col_idx)
        apply_cell_style(cell, header, font_type="data", bg_color="FFCCE5FF")

    current_row = header_row + 1

    for region_data in security_services:
        region = region_data.get("region", "")
        if region == "global":
            continue

        data = region_data.get("data", {})
        service_entries = list(data.items())

        if not service_entries:
            continue

        start = current_row
        end = current_row + len(service_entries) - 1

        # Merge region cells
        if start < end:
            sheet.merge_cells(f"B{start}:B{end}")

        for idx, (service_name, service_result) in enumerate(service_entries):
            row = current_row + idx

            if idx == 0:
                apply_cell_style(sheet.cell(row=row, column=2), region)

            is_enabled = service_result.get("is_enabled", "N/A")
            recommendation = (
                "-"
                if is_enabled.lower() == "yes"
                else service_result.get("recommendation", "N/A")
            )

            apply_cell_style(sheet.cell(row=row, column=3), service_name)
            apply_cell_style(sheet.cell(row=row, column=4), is_enabled)
            apply_cell_style(sheet.cell(row=row, column=5), recommendation)

        current_row += len(service_entries)


# ============================================================================
# REGION WORKSHEETS GENERATION
# ============================================================================


def create_region_worksheets(
    workbook: Workbook, filtered_data: List[Dict], meta: Dict
) -> None:
    """Create worksheets for each region with detailed findings"""
    # Group data by region
    region_map = {}
    for item in filtered_data:
        region_key = (
            "Global Services"
            if item.get("region") == "global"
            else item.get("region", "Unknown Region")
        )
        if region_key not in region_map:
            region_map[region_key] = []
        region_map[region_key].append(item)

    # Create sheet for each region
    for region, region_data in region_map.items():
        region_data = sort_by_severity(region_data)
        create_region_sheet(workbook, region, region_data, meta)


def create_region_sheet(
    workbook: Workbook, region: str, region_data: List[Dict], meta: Dict
) -> None:
    """Create a single region worksheet with findings"""
    sheet = workbook.create_sheet(region)

    # Set up headers
    headers = [
        {"cell": "A1", "text": "Check Name"},
        {"cell": "B1", "text": "Problem Statement"},
        {"cell": "C1", "text": "Severity Level"},
        {"cell": "D1", "text": "Region"},
        {"cell": "E1", "text": "Account ID"},
    ]

    for header in headers:
        cell = sheet[header["cell"]]
        apply_cell_style(cell, header["text"], font_type="header", bg_color="FFFFFF")

    # Set column widths
    widths = [30, 50, 15, 15, 20]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    # Calculate max resource columns
    max_resource_columns = calculate_max_resource_columns(region_data)

    # Add resource header
    if max_resource_columns > 0:
        resource_start_col = 6
        resource_end_col = resource_start_col + max_resource_columns - 1
        start_letter = get_column_letter(resource_start_col)
        end_letter = get_column_letter(resource_end_col)

        sheet.merge_cells(f"{start_letter}1:{end_letter}1")
        resource_header = sheet[f"{start_letter}1"]
        apply_cell_style(
            resource_header, "Resource Details", font_type="header", bg_color="FFFFFF"
        )

    current_row = 2

    # Add findings
    for item in region_data:
        current_row = add_finding_to_sheet(sheet, item, current_row)

    # Freeze first row
    sheet.freeze_panes = "A2"


def calculate_max_resource_columns(region_data: List[Dict]) -> int:
    """Calculate maximum number of resource columns needed"""
    max_cols = 0
    for item in region_data:
        resources = item.get("resources_affected", [])
        for res in resources:
            if res:
                max_cols = max(max_cols, len(res.keys()))
    return max_cols


def add_finding_to_sheet(sheet, item: Dict, current_row: int) -> int:
    """Add a single finding with its resources to the sheet"""
    resource_list = item.get("resources_affected", [])
    severity_color = get_severity_color(item.get("severity_level", ""))
    resource_start_col = 6

    # Add check name heading
    if item.get("check_name"):
        sheet.merge_cells(f"A{current_row}:B{current_row}")
        heading_cell = sheet[f"A{current_row}"]
        apply_cell_style(
            heading_cell,
            item["check_name"],
            font_type="resource-header",
            bg_color="FFEFEFEF",
        )
        current_row += 1

    # Add additional info
    if item.get("additional_info") and isinstance(item["additional_info"], dict):
        for key, value in item["additional_info"].items():
            formatted_key = format_key_name(key)
            apply_cell_style(
                sheet[f"A{current_row}"],
                formatted_key,
                font_type="resource-header",
                bg_color="FFD9D9D9",
            )
            apply_cell_style(
                sheet[f"B{current_row}"], value, font_type="data", bg_color="FFDDEEFF"
            )
            current_row += 1

    # Get ordered keys for resources
    ordered_keys = get_ordered_resource_keys(resource_list)

    # Create resource headers
    if ordered_keys:
        current_row = add_resource_headers(
            sheet, ordered_keys, current_row, resource_start_col, severity_color
        )

    # Add resource data
    current_row = add_resource_data(
        sheet,
        resource_list,
        ordered_keys,
        item,
        current_row,
        resource_start_col,
        severity_color,
    )

    return current_row


def format_key_name(key: str) -> str:
    """Format a key name for display"""
    if key.lower() == "total_scanned":
        return "Total no. of Resources"
    elif key.lower() == "affected":
        return "No. of Issue found"
    else:
        return key.replace("_", " ").title()


def get_ordered_resource_keys(resource_list: List[Dict]) -> List[Dict]:
    """Get ordered keys from resource list, handling nested structures"""
    if not resource_list:
        return []

    ordered_keys = []
    first_resource = resource_list[0] if resource_list else {}

    for key in first_resource.keys():
        val = first_resource[key]
        if isinstance(val, list) and val and isinstance(val[0], dict):
            # Nested structure
            sub_keys = list(
                set(
                    sub_key
                    for res in resource_list
                    for item in res.get(key, [])
                    for sub_key in (item.keys() if isinstance(item, dict) else [])
                )
            )
            ordered_keys.append({"type": "nested", "key": key, "sub_keys": sub_keys})
        else:
            ordered_keys.append({"type": "flat", "key": key})

    return ordered_keys


def add_resource_headers(
    sheet,
    ordered_keys: List[Dict],
    current_row: int,
    start_col: int,
    severity_color: str,
) -> int:
    """Add two-row headers for resource details"""
    current_col = start_col

    for key_info in ordered_keys:
        if key_info["type"] == "nested":
            col_start = current_col

            # Add sub-headers
            for sub_key in key_info["sub_keys"]:
                cell = sheet.cell(row=current_row + 1, column=current_col)
                apply_cell_style(
                    cell, sub_key, font_type="resource-header", bg_color=severity_color
                )
                sheet.column_dimensions[get_column_letter(current_col)].width = 25
                current_col += 1

            # Merge parent header
            if col_start < current_col - 1:
                start_letter = get_column_letter(col_start)
                end_letter = get_column_letter(current_col - 1)
                sheet.merge_cells(
                    f"{start_letter}{current_row}:{end_letter}{current_row}"
                )

            parent_cell = sheet.cell(row=current_row, column=col_start)
            formatted_key = key_info["key"].replace("_", " ").title()
            apply_cell_style(
                parent_cell,
                formatted_key,
                font_type="resource-header",
                bg_color=severity_color,
            )
        else:
            # Flat key - merge vertically
            col_letter = get_column_letter(current_col)
            sheet.merge_cells(
                f"{col_letter}{current_row}:{col_letter}{current_row + 1}"
            )

            cell = sheet.cell(row=current_row, column=current_col)
            apply_cell_style(
                cell,
                key_info["key"],
                font_type="resource-header",
                bg_color=severity_color,
            )

            # Apply border to merged bottom cell
            sheet.cell(row=current_row + 1, column=current_col).border = (
                get_border_style()
            )
            sheet.column_dimensions[get_column_letter(current_col)].width = 25
            current_col += 1

    return current_row + 2


def add_resource_data(
    sheet,
    resource_list: List[Dict],
    ordered_keys: List[Dict],
    item: Dict,
    current_row: int,
    start_col: int,
    severity_color: str,
) -> int:
    """Add resource data rows"""
    for res in resource_list:
        # Calculate nested rows count
        nested_rows = max(
            1,
            *(
                len(res.get(key_info["key"], []))
                for key_info in ordered_keys
                if key_info["type"] == "nested"
            ),
        )

        row_start = current_row

        for i in range(nested_rows):
            col_index = start_col

            # Add resource columns
            for key_info in ordered_keys:
                if key_info["type"] == "nested":
                    nested_item = (
                        res.get(key_info["key"], [])[i]
                        if i < len(res.get(key_info["key"], []))
                        else {}
                    )
                    for sub_key in key_info["sub_keys"]:
                        val = (
                            nested_item.get(sub_key)
                            if isinstance(nested_item, dict)
                            else None
                        )
                        cell = sheet.cell(row=current_row, column=col_index)
                        display_val = "-" if val in (None, "") else val
                        apply_cell_style(
                            cell, display_val, font_type="data", bg_color=severity_color
                        )
                        col_index += 1
                else:
                    cell = sheet.cell(row=current_row, column=col_index)
                    val = res.get(key_info["key"])
                    display_val = "-" if val in (None, "") else val
                    apply_cell_style(
                        cell, display_val, font_type="data", bg_color=severity_color
                    )
                    col_index += 1

            # Add main columns on first row only
            if i == 0:
                columns = ["A", "B", "C", "D", "E"]
                values = [
                    item.get("check_name"),
                    item.get("problem_statement"),
                    item.get("severity_level"),
                    item.get("region"),
                    item.get("account_id"),
                ]
                for col, val in zip(columns, values):
                    cell = sheet[f"{col}{current_row}"]
                    apply_cell_style(
                        cell, val, font_type="data", bg_color=severity_color
                    )

            current_row += 1

        # Merge main columns if multiple nested rows
        if nested_rows > 1:
            for col in ["A", "B", "C", "D", "E"]:
                sheet.merge_cells(f"{col}{row_start}:{col}{current_row - 1}")

    return current_row


# ============================================================================
# MAIN EXPORT FUNCTION
# ============================================================================


def generate_excel_report(data):
    """
    Generate complete Excel report with summary and region worksheets

    Args:
        meta: Metadata about the account and scan
        filtered_data: List of findings/issues
        security_services: Security services scan results

    Returns:
        BytesIO: Excel file as bytes
    """
    workbook = Workbook()

    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Create summary sheet
    create_summary_sheet(workbook, data)

    # Create region worksheets
    create_region_worksheets(workbook, data)

    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


# ============================================================================
# FASTAPI ENDPOINTS
# ============================================================================


from utils.upload_to_s3 import get_report_from_s3_function
from Model.model import ReportRequest
from fastapi.responses import Response


def get_report_word_function(requestData: ReportRequest):
    """
    Generate AWS Security Findings Report in Word format
    """
    global SECTION_NO
    SECTION_NO = 0

    response = get_report_from_s3_function(requestData)

    # Check response status
    if response.get("status", "") == "error":
        return response

    data = response.get("data", {})

    excel_file = generate_excel_report(data=data)

    # === Response Headers ===
    account_id = requestData.account_id or ""
    account_name = data.get("account_name", "")
    filename = ""

    # Case 1: Both account_id and account_name exist
    if account_id and account_name:
        filename = f"{account_id} ({account_name})_AWS_Security_Assessment_Report.docx"

    # Case 2: Only account_id exists
    elif account_id:
        filename = f"{account_id}_AWS_Security_Assessment_Report.docx"

    # Case 3: Neither exists
    else:
        filename = "AWS_Security_Assessment_Report.docx"

    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }

    return Response(
        content=excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )